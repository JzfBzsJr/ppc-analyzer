from http.server import BaseHTTPRequestHandler
import io
import json
import cgi
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Helpers ──────────────────────────────────────────────────────────────────

def safe_div(a, b, default=0):
    if b == 0 or (isinstance(b, float) and np.isnan(b)):
        return default
    return a / b


def extract_campaign_type(name):
    u = name.upper()
    if "AUTO|" in u or "AUTO " in u or u.startswith("AUTO"):
        return "Auto"
    if "PAT|" in u:
        return "Product Targeting"
    if "EXACT|" in u:
        return "Exact Match"
    if "BROAD|" in u:
        return "Broad Match"
    if "PHRASE|" in u:
        return "Phrase Match"
    if "NICHE" in u:
        return "Niche/AMZ Audience"
    return "Other"


def analyze(file_bytes):
    df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = df.columns.str.strip()

    required = {"Campaign Name", "Date", "Impressions", "Clicks", "Spend",
                "7 Day Total Sales", "7 Day Total Orders (#)"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Отчёт не содержит нужных колонок: {missing}")

    df["Campaign Type"] = df["Campaign Name"].apply(extract_campaign_type)

    total_spend      = df["Spend"].sum()
    total_sales      = df["7 Day Total Sales"].sum()
    total_orders     = df["7 Day Total Orders (#)"].sum()
    total_units      = df.get("7 Day Total Units (#)", pd.Series([0])).sum()
    total_impressions = df["Impressions"].sum()
    total_clicks     = df["Clicks"].sum()

    overall = {
        "date_start":   str(df["Date"].min().date()),
        "date_end":     str(df["Date"].max().date()),
        "days":         (df["Date"].max() - df["Date"].min()).days + 1,
        "spend":        round(total_spend, 2),
        "sales":        round(total_sales, 2),
        "orders":       int(total_orders),
        "units":        int(total_units),
        "impressions":  int(total_impressions),
        "clicks":       int(total_clicks),
        "acos":         round(safe_div(total_spend, total_sales) * 100, 2),
        "roas":         round(safe_div(total_sales, total_spend), 2),
        "ctr":          round(safe_div(total_clicks, total_impressions) * 100, 2),
        "cvr":          round(safe_div(total_orders, total_clicks) * 100, 2),
        "cpc":          round(safe_div(total_spend, total_clicks), 2),
        "aov":          round(safe_div(total_sales, total_orders), 2),
    }

    # Products
    asin_cols = {"Impressions", "Clicks", "Spend", "7 Day Total Sales", "7 Day Total Orders (#)"}
    if "7 Day Advertised SKU Sales" in df.columns:
        asin_cols.add("7 Day Advertised SKU Sales")
    if "7 Day Other SKU Sales" in df.columns:
        asin_cols.add("7 Day Other SKU Sales")

    group_keys = ["Advertised ASIN", "Advertised SKU"] if "Advertised ASIN" in df.columns else ["Campaign Name"]
    asin_perf = df.groupby(group_keys)[list(asin_cols)].sum().reset_index()

    products = []
    for _, r in asin_perf.iterrows():
        if r["Spend"] > 0:
            halo = round(r.get("7 Day Other SKU Sales", 0), 2)
            products.append({
                "asin":       r.get("Advertised ASIN", ""),
                "sku":        r.get("Advertised SKU", r.get("Campaign Name", "")),
                "spend":      round(r["Spend"], 2),
                "sales":      round(r["7 Day Total Sales"], 2),
                "orders":     int(r["7 Day Total Orders (#)"]),
                "impressions": int(r["Impressions"]),
                "clicks":     int(r["Clicks"]),
                "acos":       round(safe_div(r["Spend"], r["7 Day Total Sales"]) * 100, 2),
                "roas":       round(safe_div(r["7 Day Total Sales"], r["Spend"]), 2),
                "ctr":        round(safe_div(r["Clicks"], r["Impressions"]) * 100, 2),
                "cvr":        round(safe_div(r["7 Day Total Orders (#)"], r["Clicks"]) * 100, 2),
                "cpc":        round(safe_div(r["Spend"], r["Clicks"]), 2),
                "halo_sales": halo,
            })

    # Campaign types
    type_perf = df.groupby("Campaign Type").agg({
        "Impressions": "sum", "Clicks": "sum", "Spend": "sum",
        "7 Day Total Sales": "sum", "7 Day Total Orders (#)": "sum"
    }).reset_index()

    campaign_types = []
    for _, r in type_perf.iterrows():
        campaign_types.append({
            "type":   r["Campaign Type"],
            "spend":  round(r["Spend"], 2),
            "sales":  round(r["7 Day Total Sales"], 2),
            "orders": int(r["7 Day Total Orders (#)"]),
            "acos":   round(safe_div(r["Spend"], r["7 Day Total Sales"]) * 100, 2),
            "cvr":    round(safe_div(r["7 Day Total Orders (#)"], r["Clicks"]) * 100, 2),
            "cpc":    round(safe_div(r["Spend"], r["Clicks"]), 2),
        })

    # Campaign-level
    camp_perf = df.groupby(["Campaign Name"]).agg({
        "Impressions": "sum", "Clicks": "sum", "Spend": "sum",
        "7 Day Total Sales": "sum", "7 Day Total Orders (#)": "sum"
    }).reset_index()

    # Top performers
    profitable = camp_perf[(camp_perf["7 Day Total Sales"] > 0) & (camp_perf["Spend"] > 5)].copy()
    profitable["ACOS"] = profitable["Spend"] / profitable["7 Day Total Sales"] * 100
    profitable["CVR"]  = profitable["7 Day Total Orders (#)"] / profitable["Clicks"].replace(0, np.nan) * 100
    profitable = profitable.sort_values("ACOS").head(15)

    top_performers = []
    for _, r in profitable.iterrows():
        top_performers.append({
            "campaign": r["Campaign Name"],
            "spend":    round(r["Spend"], 2),
            "sales":    round(r["7 Day Total Sales"], 2),
            "orders":   int(r["7 Day Total Orders (#)"]),
            "acos":     round(r["ACOS"], 2),
            "cvr":      round(r["CVR"] if not np.isnan(r["CVR"]) else 0, 2),
        })

    # Wasted spend
    wasters = camp_perf[(camp_perf["Spend"] > 5) & (camp_perf["7 Day Total Sales"] == 0)].sort_values("Spend", ascending=False)
    wasted_spend = []
    for _, r in wasters.iterrows():
        wasted_spend.append({
            "campaign": r["Campaign Name"],
            "spend":    round(r["Spend"], 2),
            "clicks":   int(r["Clicks"]),
        })

    # Daily trends
    daily = df.groupby("Date").agg({
        "Impressions": "sum", "Clicks": "sum", "Spend": "sum",
        "7 Day Total Sales": "sum", "7 Day Total Orders (#)": "sum"
    }).reset_index()

    daily_trends = []
    for _, r in daily.iterrows():
        daily_trends.append({
            "date":   str(r["Date"].date()),
            "spend":  round(r["Spend"], 2),
            "sales":  round(r["7 Day Total Sales"], 2),
            "orders": int(r["7 Day Total Orders (#)"]),
            "acos":   round(safe_div(r["Spend"], r["7 Day Total Sales"]) * 100, 2),
        })

    return {
        "overall": overall,
        "products": products,
        "campaign_types": campaign_types,
        "top_performers": top_performers,
        "wasted_spend": wasted_spend,
        "daily_trends": daily_trends,
    }


# ── Excel generation ──────────────────────────────────────────────────────────

GREEN  = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED    = PatternFill("solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD   = Font(bold=True)

thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def acos_fill(v):
    if v == 0 or v > 999:
        return None
    if v < 25:
        return GREEN
    if v < 40:
        return YELLOW
    return RED


def cvr_fill(v):
    if v >= 10:
        return GREEN
    if v >= 5:
        return YELLOW
    return RED


def style_header(ws, row_num, cols):
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER


def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)


def build_excel(data):
    wb = openpyxl.Workbook()
    o = data["overall"]

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Amazon PPC Analysis"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Period: {o['date_start']} — {o['date_end']}  ({o['days']} days)"])
    ws.append([])

    style_header(ws, 4, ["Metric", "Value", "Benchmark"])
    rows = [
        ("Total Spend",       f"${o['spend']:,.2f}",   ""),
        ("Total Sales",       f"${o['sales']:,.2f}",   ""),
        ("Orders",            o["orders"],              ""),
        ("Units",             o["units"],               ""),
        ("Impressions",       f"{o['impressions']:,}",  ""),
        ("Clicks",            f"{o['clicks']:,}",       ""),
        ("ACOS",              f"{o['acos']:.1f}%",      "Good <25%  |  Warn 25-40%  |  Bad >40%"),
        ("ROAS",              f"{o['roas']:.2f}x",      ""),
        ("CTR",               f"{o['ctr']:.2f}%",       "Good >1%  |  Warn 0.5-1%"),
        ("CVR",               f"{o['cvr']:.1f}%",       "Good >10%  |  Warn 5-10%"),
        ("Avg CPC",           f"${o['cpc']:.2f}",       ""),
        ("Avg Order Value",   f"${o['aov']:.2f}",       ""),
    ]
    for i, (metric, val, bench) in enumerate(rows, 5):
        ws.append([metric, val, bench])
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = BORDER
        if metric == "ACOS":
            fill = acos_fill(o["acos"])
            if fill:
                ws.cell(row=i, column=2).fill = fill
        if metric == "CVR":
            fill = cvr_fill(o["cvr"])
            if fill:
                ws.cell(row=i, column=2).fill = fill

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 35

    # ── Sheet 2: Products ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Products")
    headers = ["ASIN", "SKU", "Spend", "Sales", "Orders", "Impressions",
               "Clicks", "ACOS %", "ROAS", "CTR %", "CVR %", "CPC", "Halo Sales"]
    style_header(ws2, 1, headers)
    for i, p in enumerate(sorted(data["products"], key=lambda x: x["spend"], reverse=True), 2):
        row = [p["asin"], p["sku"], p["spend"], p["sales"], p["orders"],
               p["impressions"], p["clicks"], p["acos"], p["roas"],
               p["ctr"], p["cvr"], p["cpc"], p["halo_sales"]]
        ws2.append(row)
        for col in range(1, len(headers) + 1):
            ws2.cell(row=i, column=col).border = BORDER
        acos_col = ws2.cell(row=i, column=8)
        fill = acos_fill(p["acos"])
        if fill:
            acos_col.fill = fill
        cvr_col = ws2.cell(row=i, column=11)
        fill2 = cvr_fill(p["cvr"])
        if fill2:
            cvr_col.fill = fill2
    autofit(ws2)

    # ── Sheet 3: Campaign Types ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Campaign Types")
    headers3 = ["Campaign Type", "Spend", "Sales", "Orders", "ACOS %", "CVR %", "CPC", "Assessment"]
    style_header(ws3, 1, headers3)
    sorted_types = sorted(data["campaign_types"], key=lambda x: x["acos"] if x["acos"] < 999 else 999)
    for i, ct in enumerate(sorted_types, 2):
        if ct["acos"] < 25:
            assessment = "Strong - Scale"
        elif ct["acos"] < 40:
            assessment = "Monitor closely"
        elif ct["acos"] < 60:
            assessment = "Needs optimization"
        else:
            assessment = "Review / Pause"
        row = [ct["type"], ct["spend"], ct["sales"], ct["orders"],
               ct["acos"], ct["cvr"], ct["cpc"], assessment]
        ws3.append(row)
        for col in range(1, len(headers3) + 1):
            ws3.cell(row=i, column=col).border = BORDER
        fill = acos_fill(ct["acos"])
        if fill:
            ws3.cell(row=i, column=5).fill = fill
    autofit(ws3)

    # ── Sheet 4: Top Performers ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Top Performers")
    headers4 = ["Campaign", "Spend", "Sales", "Orders", "ACOS %", "CVR %", "Action"]
    style_header(ws4, 1, headers4)
    for i, p in enumerate(data["top_performers"], 2):
        if p["acos"] < 15:
            action = "Scale 3x"
        elif p["acos"] < 25:
            action = "Scale 2x"
        else:
            action = "Scale 1.5x"
        row = [p["campaign"], p["spend"], p["sales"], p["orders"], p["acos"], p["cvr"], action]
        ws4.append(row)
        for col in range(1, len(headers4) + 1):
            ws4.cell(row=i, column=col).border = BORDER
        fill = acos_fill(p["acos"])
        if fill:
            ws4.cell(row=i, column=5).fill = fill
        ws4.cell(row=i, column=7).font = Font(bold=True, color="2E5C55")
    autofit(ws4)

    # ── Sheet 5: Wasted Spend ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("Wasted Spend")
    total_wasted = sum(w["spend"] for w in data["wasted_spend"])
    ws5.append(["Campaigns with spend > $5 and ZERO sales — pause immediately"])
    ws5["A1"].font = Font(bold=True, color="CF4043")
    ws5.append([f"Total wasted: ${total_wasted:,.2f}"])
    ws5.append([])
    headers5 = ["Campaign", "Spend", "Clicks", "Action"]
    style_header(ws5, 4, headers5)
    for i, w in enumerate(data["wasted_spend"], 5):
        ws5.append([w["campaign"], w["spend"], w["clicks"], "PAUSE"])
        for col in range(1, 5):
            ws5.cell(row=i, column=col).border = BORDER
        ws5.cell(row=i, column=2).fill = RED
        ws5.cell(row=i, column=4).font = Font(bold=True, color="CF4043")
    autofit(ws5)
    ws5.column_dimensions["A"].width = 60

    # ── Sheet 6: Daily Trends ─────────────────────────────────────────────────
    ws6 = wb.create_sheet("Daily Trends")
    headers6 = ["Date", "Spend", "Sales", "Orders", "ACOS %"]
    style_header(ws6, 1, headers6)
    for i, d in enumerate(data["daily_trends"], 2):
        ws6.append([d["date"], d["spend"], d["sales"], d["orders"], d["acos"]])
        for col in range(1, 6):
            ws6.cell(row=i, column=col).border = BORDER
        fill = acos_fill(d["acos"])
        if fill:
            ws6.cell(row=i, column=5).fill = fill
    autofit(ws6)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Vercel handler ────────────────────────────────────────────────────────────

CORS_HEADERS = {
    "Access-Control-Allow-Origin":  "*",
    "Access-Control-Allow-Methods": "POST, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type",
}


class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        for k, v in CORS_HEADERS.items():
            self.send_header(k, v)
        self.end_headers()

    def do_POST(self):
        try:
            content_type = self.headers.get("Content-Type", "")
            content_length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(content_length)

            # Parse multipart
            environ = {
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE":   content_type,
                "CONTENT_LENGTH": str(content_length),
            }
            fs = cgi.FieldStorage(
                fp=io.BytesIO(body),
                environ=environ,
                keep_blank_values=True,
            )

            if "file" not in fs:
                self._error(400, "Поле 'file' не найдено в запросе.")
                return

            file_bytes = fs["file"].file.read()
            if not file_bytes:
                self._error(400, "Файл пустой.")
                return

            data = analyze(file_bytes)
            xlsx = build_excel(data)

            self.send_response(200)
            for k, v in CORS_HEADERS.items():
                self.send_header(k, v)
            self.send_header("Content-Type",
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition",
                             "attachment; filename=\"ppc-analysis.xlsx\"")
            self.send_header("Content-Length", str(len(xlsx)))
            self.end_headers()
            self.wfile.write(xlsx)

        except ValueError as e:
            self._error(400, str(e))
        except Exception as e:
            self._error(500, f"Внутренняя ошибка: {e}")

    def _error(self, code, message):
        body = json.dumps({"error": message}).encode()
        self.send_response(code)
        for k, v in CORS_HEADERS.items():
            self.send_header(k, v)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, *args):
        pass  # suppress default Vercel logs
